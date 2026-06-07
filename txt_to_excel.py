#!/usr/bin/env python3
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def build_advanced_excel_report():
    print("[+] Initializing Ultra-Fast Excel Reporter Engine...")
    
    # 1. 마스터 타깃 목록 로드
    if not os.path.exists('targets.txt'):
        print("[-] Error: targets.txt missing.")
        return
        
    with open('targets.txt', 'r') as f:
        targets = [line.strip() for line in f if line.strip()]

    # 도메인별 데이터 구조화
    matrix_data = {domain: set() for domain in targets}

    # 중복 파일 탐색 방지 (recursive=True 만으로 상하위 폴더 깔끔하게 수집)
    txt_files = set(glob.glob('results/**/*', recursive=True))
    txt_files = [f for f in txt_files if os.path.isfile(f)]

    print(f"[+] Processing {len(txt_files)} data source files...")
    
    for file_path in txt_files:
        filename = os.path.basename(file_path).lower()
        
        if 'secretfinder' in filename:
            source_tool = 'SecretFinder'
        elif 'waybackurls' in filename:
            source_tool = 'Waybackurls'
        elif 'gau' in filename:
            source_tool = 'GAU'
        else:
            source_tool = 'Combined-Engine'

        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    url = line.strip()
                    if not url or url.startswith('#'):
                        continue
                    
                    # URL 매핑 최적화
                    for domain in targets:
                        if domain in url:
                            matrix_data[domain].add((url, source_tool))
                            break
        except Exception as e:
            print(f"[-] Error reading {filename}: {e}")

    # 3. 고품격 엑셀 문서 디자인 빌드
    wb = Workbook()
    default_sheet = wb.active

    font_header = Font(name='Malgun Gothic', size=11, bold=True, color='FFFFFF')
    font_body = Font(name='Malgun Gothic', size=10, bold=False)
    fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    sheets_created = 0

    for domain, dataset in matrix_data.items():
        if not dataset:
            continue
            
        safe_tab_name = domain[:30]
        ws = wb.create_sheet(title=safe_tab_name)
        sheets_created += 1

        headers = ["No", "Target URL / Endpoint (수집된 자산 주소)", "Source Tool (발견 도구)"]
        ws.append(headers)

        ws.row_dimensions[1].height = 26
        for col_num, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        sorted_dataset = sorted(list(dataset), key=lambda x: (x[1], x[0]))
        
        # [★초고속 최적화 핵심★] 실시간 글자 길이를 저장할 변수 초기화
        max_len_no = len("No")
        max_len_url = len("Target URL / Endpoint (수집된 자산 주소)")
        max_len_tool = len("Source Tool (발견 도구)")
        
        for idx, (url, tool) in enumerate(sorted_dataset, 1):
            # 엑셀 시트당 최대 행 제한(1,048,576) 오버플로우 방지 안전장치
            if idx > 1048500:
                print(f"[!] Warning: {domain} data truncated due to Excel row limit.")
                break
                
            ws.append([idx, url, tool])
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20
            
            # 스타일 주입
            ws.cell(row=current_row, column=1).font = font_body
            ws.cell(row=current_row, column=1).alignment = align_center
            
            ws.cell(row=current_row, column=2).font = font_body
            ws.cell(row=current_row, column=2).alignment = align_left
            
            ws.cell(row=current_row, column=3).font = font_body
            ws.cell(row=current_row, column=3).alignment = align_center

            # [★초고속 최적화 핵심★] 셀을 적으면서 동시에 최대 길이를 실시간 업데이트
            max_len_no = max(max_len_no, len(str(idx)))
            max_len_url = max(max_len_url, len(str(url)))
            max_len_tool = max(max_len_tool, len(str(tool)))

        # 자동 필터 영역 적용
        max_row = ws.max_row
        ws.auto_filter.ref = f"A1:C{max_row}"

        # 루프를 다시 돌지 않고, 실시간 계측된 값으로 컬럼 너비 즉시 다이렉트 주입
        ws.column_dimensions['A'].width = max_len_no + 3
        ws.column_dimensions['B'].width = max(min(max_len_url + 3, 90), 10)
        ws.column_dimensions['C'].width = max_len_tool + 3

    # 4. 저장
    if sheets_created > 0:
        wb.remove(default_sheet)
        os.makedirs('reports', exist_ok=True)
        report_path = 'reports/passive_recon_report_v1.xlsx'
        wb.save(report_path)
        print(f"[+] [SUCCESS] Advanced filtered report generated at: {report_path}")
    else:
        print("[-] Error: Scan results were empty. Excel file not created.")

if __name__ == '__main__':
    build_advanced_excel_report()
